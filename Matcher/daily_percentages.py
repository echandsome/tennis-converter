import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def process_file():
    file1_path = file1_entry.get()
    file2_path = file2_entry.get()

    if not file1_path or not file2_path:
        result_label.config(text="Please select both Excel files.")
        return

    try:
        df_hr = pd.read_excel(file1_path, engine="openpyxl", usecols=[0, 3], header=None)
        df_daily = pd.read_excel(file2_path, engine="openpyxl", header=None)

        # (A, D) pair → key set
        hr_keys = set(tuple(row) for row in df_hr.values)

        # Load Excel file
        wb = load_workbook(file2_path)
        ws = wb.active

        # Define background color (green shade)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Apply background color to rows with matching keys
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
            key = (row[0].value, row[13].value)
            if key in hr_keys:
                for cell in row:
                    cell.fill = green_fill

        # Save
        original_filename = os.path.basename(file2_path)
        output_filename = f"Result_{original_filename}"
        output_path = os.path.join(os.path.dirname(file2_path), output_filename)
        wb.save(output_path)

        result_label.config(text=f"Highlighting complete! Saved at:\n{output_path}")
    except Exception as e:
        result_label.config(text=f"Error: {str(e)}")

def browse_file1():
    filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    file1_entry.delete(0, tk.END)
    file1_entry.insert(0, filename)

def browse_file2():
    filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    file2_entry.delete(0, tk.END)
    file2_entry.insert(0, filename)

# GUI Setup
root = tk.Tk()
root.title("Excel Row Highlighter (Green Background)")
root.geometry("480x300")

tk.Label(root, text="Select 'Percentages.xlsx':").pack()
file1_entry = tk.Entry(root, width=60)
file1_entry.pack()
tk.Button(root, text="Browse", command=browse_file1).pack(pady=5)

tk.Label(root, text="Select 'Daily_Home.xlsx':").pack()
file2_entry = tk.Entry(root, width=60)
file2_entry.pack()
tk.Button(root, text="Browse", command=browse_file2).pack(pady=5)

tk.Button(root, text="Process & Highlight Matches", command=process_file).pack(pady=15)

result_label = tk.Label(root, text="", wraplength=460)
result_label.pack()

root.mainloop()
