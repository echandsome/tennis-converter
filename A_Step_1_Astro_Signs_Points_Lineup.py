import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime
import re

def convert_date(date_str):
    """Convert 'DD_MMM_YYYY' to 'M/D/YYYY' format."""
    if pd.isna(date_str):
        return None
    
    clean_date = re.sub(r"[()]", "", date_str.strip())

    parts = clean_date.split("_")
    if len(parts) == 3:
        day, month, year = parts
        try:
            converted_date = datetime.strptime(f"{day} {month} {year}", "%d %b %Y").strftime("%m/%d/%Y")
            return converted_date.lstrip("0").replace("/0", "/")
        except ValueError:
            return None
    return None

def process_csv_to_excel(csv_file, excel_file, reference_excel):
    ref_wb = load_workbook(reference_excel, data_only=True)
    ref_ws = ref_wb.active
    df = pd.read_csv(csv_file, header=None)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Partner A", "", "Partner B", ""] + [""] * 7 + ["Signs", "Symbol"])
    
    ref_col_indices = [53, 64, 54, 69, 62, 73, 74, 75, 65, 67, 68]
    
    for _, row in df.iterrows():
        if pd.notna(row[0]) and pd.notna(row[2]):
            dob_converted = convert_date(row[2]) if pd.notna(row[2]) else None
            ws.append([row[0], row[1], row[2], row[3]] + [""] * 7)
            if dob_converted:
                for ref_row in ref_ws.iter_rows(min_row=2, max_row=ref_ws.max_row, min_col=64, max_col=64):
                    if str(ref_row[0].value).strip() == dob_converted:
                        extracted_data = [ref_ws.cell(row=ref_row[0].row, column=col).value for col in ref_col_indices]
                        ws.append(extracted_data)
        elif pd.notna(row[4]):
            ws.append([""] * 11 + [row[4], row[5]])
    
    wb.save(excel_file)
    messagebox.showinfo("Success", "Excel file saved successfully!")

def browse_csv():
    file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    csv_entry.delete(0, tk.END)
    csv_entry.insert(0, file_path)

def browse_excel():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    excel_entry.delete(0, tk.END)
    excel_entry.insert(0, file_path)

def run_process():
    csv_file = csv_entry.get()
    ref_file = excel_entry.get()
    if not csv_file or not ref_file:
        messagebox.showerror("Error", "Please select both CSV and Excel files.")
        return
    process_csv_to_excel(csv_file, "output.xlsx", ref_file)

# Tkinter GUI Setup
root = tk.Tk()
root.title("CSV to Excel Processor")
root.geometry("500x250")

# CSV File Selection
tk.Label(root, text="Select CSV File:").pack(pady=5)
csv_entry = tk.Entry(root, width=50)
csv_entry.pack()
tk.Button(root, text="Browse", command=browse_csv).pack()

# Excel File Selection
tk.Label(root, text="Select Reference Excel File:").pack(pady=5)
excel_entry = tk.Entry(root, width=50)
excel_entry.pack()
tk.Button(root, text="Browse", command=browse_excel).pack()

# Process Button
tk.Button(root, text="Process", command=run_process, bg="green", fg="white").pack(pady=10)

root.mainloop()