import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import os

class ExcelProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Filler")
        self.root.geometry("300x150")
        
        self.button = tk.Button(root, text="Select Excel File", command=self.process_file)
        self.button.pack(pady=20)
    
    def get_filename(self, file_path):
        return os.path.basename(file_path)
    
    def process_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        
        origin_file_name = self.get_filename(file_path)
        output_file_name = f'Result____{origin_file_name}.xlsx'
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            ws[f"N1"] = 'Signs-Symbol'
            
            for row in range(2, ws.max_row + 1):
                l_value = ws[f"L{row}"].value or ""
                m_value = ws[f"M{row}"].value or ""
                ws[f"N{row}"] = f"{l_value}-{m_value}" if l_value or m_value else ""

            for row in range(2, ws.max_row):  
                ws[f"L{row}"] = ws[f"L{row + 1}"].value
                ws[f"M{row}"] = ws[f"M{row + 1}"].value
                ws[f"N{row}"] = ws[f"N{row + 1}"].value

            ws[f"L{ws.max_row}"] = None
            ws[f"M{ws.max_row}"] = None
            ws[f"N{ws.max_row}"] = None

            columns_to_fill = [chr(col) for col in range(ord('A'), ord('K') + 1)]
            last_filled_values = {col: None for col in columns_to_fill}

            for row in range(2, ws.max_row):
                if all(ws[f"A{row}"].value is None for col in columns_to_fill) and ws[f"L{row}"].value is not None:
                    for col in columns_to_fill:
                        ws[f"{col}{row}"] = last_filled_values[col]
                else:
                    for col in columns_to_fill:
                        last_filled_values[col] = ws[f"{col}{row}"].value
            
            output_path = os.path.join(os.path.dirname(file_path), output_file_name)
            wb.save(output_path)
            messagebox.showinfo("Success", f"File saved as: {output_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessor(root)
    root.mainloop()
