import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime
import re
import openpyxl
import os
import tempfile

TEMP_DIR = tempfile.mkdtemp()

def convert_date(date_str):
    """Convert 'DD_MMM_YYYY' to 'M/D/YYYY' format."""
    if pd.isna(date_str):
        return None
    
    clean_date = re.sub(r"[()]", "", date_str.strip())

    parts = clean_date.split("_")
    if len(parts) == 3:
        day, month, year = parts
        try:
            converted_date = datetime.strptime(f"{day} {month} {year}", "%d %b %Y").strftime("%m/%d/%Y")
            return converted_date.lstrip("0").replace("/0", "/")
        except ValueError:
            return None
    return None

def step1_signs_points_lineup(csv_file, ref_file):
    output_path = os.path.join(TEMP_DIR, "output.xlsx")
    ref_wb = load_workbook(ref_file, data_only=True)
    ref_ws = ref_wb.active
    df = pd.read_csv(csv_file, header=None)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Partner A", "", "Partner B", ""] + [""] * 7 + ["Signs", "Symbol"])
    
    ref_col_indices = [53, 64, 54, 69, 62, 73, 74, 75, 65, 67, 68]
    
    ref_row_num = 2
    
    for _, row in df.iterrows():
        if pd.notna(row[0]) and pd.notna(row[2]):
            dob_converted = convert_date(row[2]) if pd.notna(row[2]) else None
            ws.append([row[0], row[1], row[2], row[3]] + [""] * 7)
            if dob_converted:
                if ref_row_num <= ref_ws.max_row:
                    extracted_data = [
                        ref_ws.cell(row=ref_row_num, column=col).value for col in ref_col_indices
                    ]
                    ws.append(extracted_data)
                    ref_row_num += 1
        elif pd.notna(row[4]):
            ws.append([""] * 11 + [row[4], row[5]])
    
    wb.save(output_path)
    step2_filter_fl(output_path, ref_file)
    print("Success", "Step1 file saved successfully!")

def step2_filter_fl(input_path, output_path):
    
    origin_file_name = os.path.basename(input_path)
    output_file_name = f'Result____{origin_file_name}.xlsx'
    try:
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active

        ws[f"N1"] = 'Signs-Symbol'
        
        for row in range(2, ws.max_row + 1):
            l_value = ws[f"L{row}"].value or ""
            m_value = ws[f"M{row}"].value or ""
            ws[f"N{row}"] = f"{l_value}-{m_value}" if l_value or m_value else ""

        for row in range(2, ws.max_row):  
            ws[f"L{row}"] = ws[f"L{row + 1}"].value
            ws[f"M{row}"] = ws[f"M{row + 1}"].value
            ws[f"N{row}"] = ws[f"N{row + 1}"].value

        ws[f"L{ws.max_row}"] = None
        ws[f"M{ws.max_row}"] = None
        ws[f"N{ws.max_row}"] = None

        columns_to_fill = [chr(col) for col in range(ord('A'), ord('K') + 1)]
        last_filled_values = {col: None for col in columns_to_fill}

        for row in range(2, ws.max_row):
            if all(ws[f"A{row}"].value is None for col in columns_to_fill) and ws[f"L{row}"].value is not None:
                for col in columns_to_fill:
                    ws[f"{col}{row}"] = last_filled_values[col]
            else:
                for col in columns_to_fill:
                    last_filled_values[col] = ws[f"{col}{row}"].value
        
        output_path = os.path.join(os.path.dirname(output_path), output_file_name)
        wb.save(output_path)
        messagebox.showinfo("Success", f"File saved as: {output_path}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def browse_csv():
    file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    csv_entry.delete(0, tk.END)
    csv_entry.insert(0, file_path)

def browse_excel():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    excel_entry.delete(0, tk.END)
    excel_entry.insert(0, file_path)

def run_process():
    csv_file = csv_entry.get()
    ref_file = excel_entry.get()
    if not csv_file or not ref_file:
        messagebox.showerror("Error", "Please select both CSV and Excel files.")
        return
    step1_signs_points_lineup(csv_file, ref_file)

# Tkinter GUI Setup
root = tk.Tk()
root.title("CSV to Excel Processor")
root.geometry("500x250")

# CSV File Selection
tk.Label(root, text="Select CSV File:").pack(pady=5)
csv_entry = tk.Entry(root, width=50)
csv_entry.pack()
tk.Button(root, text="Browse", command=browse_csv).pack()

# Excel File Selection
tk.Label(root, text="Select Reference Excel File:").pack(pady=5)
excel_entry = tk.Entry(root, width=50)
excel_entry.pack()
tk.Button(root, text="Browse", command=browse_excel).pack()

# Process Button
tk.Button(root, text="Process", command=run_process, bg="green", fg="white").pack(pady=10)

root.mainloop()