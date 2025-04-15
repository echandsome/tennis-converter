import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import tempfile
import os

TEMP_DIR = tempfile.mkdtemp()

def browse_file(entry):
    """Open file dialog and update the entry field."""
    filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    entry.delete(0, tk.END)
    entry.insert(0, filename)

def step1_merge_files():
    """Process the input Excel files with the merging logic."""
    input_file_1 = entry1.get()
    input_file_2 = entry2.get()

    if not input_file_1 or not input_file_2:
        messagebox.showerror("Error", "Please select both input files!")
        return

    try:
        # Load the Excel files
        df1 = pd.read_excel(input_file_1, header=None)
        df2 = pd.read_excel(input_file_2, header=None)

        # Extract relevant columns from Input 1 (Columns A to G)
        df1_filtered = df1.iloc[:, [0, 1, 2, 3, 4, 5, 6]]
        df1_filtered.columns = ['Symbol', 'Phase', 'Over count', 'Under count', 'Total', 'WIN% OVER', 'WIN% UNDER']

        # Extract relevant columns from Input 2 (Columns N and Q)
        df2_filtered = df2.iloc[:, [13, 16]]
        df2_filtered.columns = ['Symbol', 'Phase']

        # Iterate through rows of Input 2 where both 'Symbol' and 'Phase' are not empty
        for index, row in df2_filtered.iterrows():
            symbol_input2 = row['Symbol']
            phase_input2 = row['Phase']

            if pd.notna(symbol_input2) and pd.notna(phase_input2):
                # Match with Input 1 where 'Symbol' and 'Phase' match
                matched_row = df1_filtered[(df1_filtered['Symbol'] == symbol_input2) & (df1_filtered['Phase'] == phase_input2)]

                if not matched_row.empty:
                    # Append matched values from Input 1 to Input 2
                    df2.loc[index, 'Over count'] = matched_row['Over count'].values[0]
                    df2.loc[index, 'Under count'] = matched_row['Under count'].values[0]
                    df2.loc[index, 'Total'] = matched_row['Total'].values[0]
                    df2.loc[index, 'WIN% OVER'] = matched_row['WIN% OVER'].values[0]
                    df2.loc[index, 'WIN% UNDER'] = matched_row['WIN% UNDER'].values[0]

                    # Check the condition for Column X and Y
                    win_percent_over = matched_row['WIN% OVER'].values[0]
                    win_percent_under = matched_row['WIN% UNDER'].values[0]

                    # Assign values to Column X and Y based on percentages
                    if win_percent_over > win_percent_under:
                        df2.loc[index, 'X'] = 'O'
                        df2.loc[index, 'Y'] = ''
                    elif win_percent_under > win_percent_over:
                        df2.loc[index, 'X'] = ''
                        df2.loc[index, 'Y'] = 'U'
                    else:  # If 50%/50%, mark both
                        df2.loc[index, 'X'] = 'O'
                        df2.loc[index, 'Y'] = 'U'

        output_path = os.path.join(TEMP_DIR, "step1_output.xlsx")
        df2.to_excel(output_path, index=False, header=False)

        # Load the saved file to modify headers
        wb = load_workbook(output_path)
        ws = wb.active

        # Add bold headers in row 1 for columns S to Y
        headers = ['Over count', 'Under count', 'Total', 'WIN% OVER', 'WIN% UNDER', 'COL VW', 'COL VW']
        header_start_col = 19  # Column S (Excel columns are 1-based)
        for i, header in enumerate(headers):
            cell = ws.cell(row=1, column=header_start_col + i, value=header)
            cell.font = Font(bold=True)  # Make the text bold

        # Save the updated file
        wb.save(output_path)

        print("Success", f"Processing completed! Output saved as {output_path}")
        step2_aggregate_players(output_path)
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


# Function to add an aggregate row to each player's DataFrame
def add_aggregate_row(df):
    print(df.columns)
    agg_row = {
        'Over count': df['Over count'].sum(),
        'Under count': df['Under count'].sum(),
        'Total': df['Total'].sum(),
        'WIN% OVER': df['WIN% OVER'].sum(),
        'WIN% UNDER': df['WIN% UNDER'].sum(),
        'COL VW': df['COL VW'].map(lambda x: x == 'O').sum(),
        'COL VW.1': df['COL VW.1'].map(lambda x: x == 'U').sum(),
        "Remarks": "Result"
    }
    
    # Copy last row to preserve other attributes
    last_row = df.iloc[-1].copy()
    for col in agg_row:
        last_row[col] = agg_row[col]
    
    # Concatenate the new row to the DataFrame
    df = pd.concat([df, last_row.to_frame().T], ignore_index=True)
    
    return df
# Function to separate each player's data into a dictionary of DataFrames
def separate_player_data(data):
    player_dfs = {}
    current_player_data = []
    player_count = 1

    for _, row in data.iterrows():
        if pd.isna(row['Partner A']) and pd.isna(row['Partner B']):
            if current_player_data:
                player_df = pd.DataFrame(current_player_data)
                player_df['Remarks'] = ""
                player_dfs[f'Player_{player_count}'] = player_df
                player_count += 1
                current_player_data = []
        else:
            current_player_data.append(row)

    if current_player_data:
        player_df = pd.DataFrame(current_player_data)
        player_df['Remarks'] = ""
        player_dfs[f'Player_{player_count}'] = player_df

    return player_dfs

# Function to handle file selection and processing
def step2_aggregate_players(input_path):

    try:
        root.update()  # Refresh the GUI
        # Load the Excel file
        data = pd.read_excel(input_path, sheet_name='Sheet1')
        original_columns = data.columns
        second_header = data.iloc[0]
        data.drop(index=0, inplace=True)
        
        # Separate player data
        player_dfs = separate_player_data(data)
        
        # Combine all player DataFrames
        combined_df = pd.DataFrame()
        for i, (player, df) in enumerate(player_dfs.items()):
            df = add_aggregate_row(df)
            if i == 0:
                combined_df = pd.concat([combined_df, df], ignore_index=True)
            else:
                combined_df = pd.concat([combined_df, df], ignore_index=True)

        # Add original headers and second row at the top
        new_columns = [None if str(x).startswith('Unnamed') else x for x in original_columns]
        if len(new_columns) < combined_df.shape[1]:
            new_columns.append("Remarks")
        combined_df.columns = new_columns
        combined_df.loc[-1] = second_header
        combined_df.index = combined_df.index + 1
        combined_df = combined_df.sort_index()

        output_path = os.path.join(TEMP_DIR, "step2_output.xlsx")
        combined_df.to_excel(output_path, index=False)

        print("Success", f"File saved to: {output_path}")
        step3_final_process(output_path)
    except Exception as e:
        raise e
        messagebox.showerror("Error", f"An error occurred: {e}")


def step3_final_process(input_path):
    try:
        
        input_filename = os.path.splitext(os.path.basename(entry2.get()))[0]
        # Create the output file name by adding "_fix" before the extension
        output_file = f"result_{input_filename}.xlsx"

        # Load the Excel file into a pandas DataFrame
        df = pd.read_excel(input_path)


        # Delete the first row (index 0)
        df = df.drop(index=0)

        # Filter rows where column Z contains "Result"
        column_z_index = 25  # Z is the 26th column, so its index is 25
        df = df[df.iloc[:, column_z_index] == 'Result']

        # Drop columns X, Y, Z (columns 24, 25, 26)
        columns_to_drop = [23, 24, 25]  # X=24, Y=25, Z=26
        df = df.drop(df.columns[columns_to_drop], axis=1)

        # Calculate column X as V - W (V is 22nd column, W is 23rd column)
        column_v_index = 21  # V is the 22nd column (index 21)
        column_w_index = 22  # W is the 23rd column (index 22)
        df['X'] = df.iloc[:, column_v_index] - df.iloc[:, column_w_index]

        # Round columns V, W, and X to 2 decimal places
        df.iloc[:, column_v_index] = df.iloc[:, column_v_index].round(2)
        df.iloc[:, column_w_index] = df.iloc[:, column_w_index].round(2)
        df['X'] = df['X'].round(2)

        # Rearrange columns and move columns B, F, G, V, W, X
        cols_bfg = df.iloc[:, [1, 5, 6]]
        cols_vwx = df.iloc[:, [21, 22, 23]]
        df = df.drop(df.columns[[1, 5, 6, 21, 22, 23]], axis=1)
        df.insert(4, 'V', cols_vwx.iloc[:, 0])
        df.insert(5, 'W', cols_vwx.iloc[:, 1])
        df.insert(6, 'X', cols_vwx.iloc[:, 2])
        df.insert(21, 'B', cols_bfg.iloc[:, 0])
        df.insert(22, 'F', cols_bfg.iloc[:, 1])
        df.insert(23, 'G', cols_bfg.iloc[:, 2])

        # Sort the DataFrame by column 'G' (new position)
        df = df.sort_values(by=df.columns[6], ascending=False)

        # Delete the contents of column I
        df.iloc[:, 8] = None  # Column I is at index 8 (9th column), so set it to None


        df.to_excel(output_file, index=False, header=False)

        # Open the saved file to apply additional formatting
        wb = load_workbook(output_file)
        ws = wb.active

        # Define colors
        green_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
        red_fill = PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # Initialize lists for green, red, and gray rows
        green_rows = []
        red_rows = []
        gray_rows = []

        # Apply conditions for coloring and sorting rows
        for row in range(1, len(df) + 2):  # Adjusting for header rows
            # Get the value in column G
            g_value = ws.cell(row=row, column=7).value

            # Skip if the value is None or NaN
            if g_value is None or pd.isna(g_value):
                continue

            # Condition 1: Color column E green if G > 0.2
            if g_value > 0.2:
                green_rows.append(row)  # Store green rows
                ws.cell(row=row, column=5).fill = green_fill  # Column E is 5
            # Condition 2: Color column E red if G <= -0.2
            elif g_value <= -0.2:
                red_rows.append(row)  # Store red rows
                ws.cell(row=row, column=6).fill = red_fill  # Column E is 5
            # Condition 3: Color columns E, F, G gray and move them to the bottom if -0.19 <= G <= 0.19
            elif -0.19 <= g_value <= 0.19:
                gray_rows.append(row)  # Store gray rows
                ws.cell(row=row, column=5).fill = gray_fill  # Column E is 5
                ws.cell(row=row, column=6).fill = gray_fill  # Column F is 6
                ws.cell(row=row, column=7).fill = gray_fill  # Column G is 7

        # Save the file with all formatting applied
        wb.save(output_file)

        # Show success message
        messagebox.showinfo("Success", f"File processed successfully! Output saved as {output_file}")

    except Exception as e:
        # Handle errors and display a message
        messagebox.showerror("Error", f"An error occurred: {e}")


# Create Tkinter GUI
root = tk.Tk()
root.title("Excel Data Merger")

# Labels
tk.Label(root, text="Select Input File 1 (1.xlsx):").grid(row=0, column=0, padx=10, pady=5, sticky="w")
tk.Label(root, text="Select Input File 2 (2.xlsx):").grid(row=1, column=0, padx=10, pady=5, sticky="w")

# Entry fields
entry1 = tk.Entry(root, width=50)
entry2 = tk.Entry(root, width=50)
entry1.grid(row=0, column=1, padx=10, pady=5)
entry2.grid(row=1, column=1, padx=10, pady=5)

# Browse buttons
btn_browse1 = tk.Button(root, text="Browse", command=lambda: browse_file(entry1))
btn_browse2 = tk.Button(root, text="Browse", command=lambda: browse_file(entry2))
btn_browse1.grid(row=0, column=2, padx=5, pady=5)
btn_browse2.grid(row=1, column=2, padx=5, pady=5)

# Process button
btn_process = tk.Button(root, text="Process", command=step1_merge_files, bg="green", fg="white", font=("Arial", 12, "bold"))
btn_process.grid(row=2, column=0, columnspan=3, pady=20)

# Run the Tkinter event loop
root.mainloop()
