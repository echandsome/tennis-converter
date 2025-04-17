import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook
import re

def get_filtered_rows(historical_path, symbol, count):
    wb = load_workbook(historical_path, data_only=True)
    ws = wb.active

    found = False
    max_count = int(count)
    data_collected = 0
    match_rows = []
    c_count = 0

    for i, row in enumerate(ws.iter_rows(min_row=1), start=1):
        b_cell = row[1]  # Column B

        if found:
            if all(cell.value is None for cell in row):
                break  # Stop when reaching an empty row
            match_rows.append(row)

        elif b_cell.value == symbol:
            found = True  # Start from the matched symbol

    for i in reversed(range(len(match_rows))):
        row = match_rows[i]

        b_val = row[1].value  # Column B
        c_val = row[2].value  # Column C

        if isinstance(b_val, (int, float)):
            c_count += b_val

        if isinstance(c_val, (int, float)):
            c_count -= c_val

        data_collected += 1
        if data_collected >= max_count:
            break

    return c_count

def process_daily_file(daily_path, historical_path, count_val):
    wb = load_workbook(daily_path)
    ws = wb.active
    total = 0
    start = False
    for i, row in enumerate(ws.iter_rows(min_row=1), start=1):
        a_value = row[0].value  # Column A
        n_value = row[13].value  # Column N

        pattern = r"^\(\d{1,2}_[A-Za-z]{1,10}_\d{4}\)$"
        if isinstance(a_value, str) and re.match(pattern, a_value.strip()):
            start = True

        if not a_value or a_value == "":
            row[10].value = total
            total = 0
            start = False
            continue

        if start == False or not a_value or not n_value:
            continue

        t_count = get_filtered_rows(historical_path, n_value, count_val)
        row[10].value = t_count  # Column K (index 10)
        total += t_count

        print(f"Row {i} ({a_value}) - Symbol: {n_value} - t_count: {t_count}")

    out_path = os.path.join(os.path.dirname(daily_path), "Modified_Daily.xlsx")
    wb.save(out_path)

    result_label.config(text=f"Completed!\nSaved to {out_path}", fg="green")

# --- GUI SETUP ---

def browse_daily():
    path = filedialog.askopenfilename(title="Select Daily.xlsx", filetypes=[("Excel files", "*.xlsx")])
    if path:
        daily_path_var.set(path)

def browse_historical():
    path = filedialog.askopenfilename(title="Select Historical Stats.xlsx", filetypes=[("Excel files", "*.xlsx")])
    if path:
        historical_path_var.set(path)

def run_conversion():
    daily_path = daily_path_var.get()
    historical_path = historical_path_var.get()
    count_val = count_entry.get()

    if not daily_path or not historical_path or not count_val:
        result_label.config(text="Please select all files and input a count value.", fg="red")
        return

    result_label.config(text="Processing...", fg="blue")
    root.update()

    try:
        process_daily_file(daily_path, historical_path, count_val)
    except Exception as e:
        print(e)
        result_label.config(text=f"Error: {str(e)}", fg="red")

root = tk.Tk()
root.title("Daily.xlsx Processor")
root.geometry("600x300")

daily_path_var = tk.StringVar()
historical_path_var = tk.StringVar()

tk.Button(root, text="Select Daily.xlsx", command=browse_daily).pack(pady=(10, 0))
tk.Label(root, textvariable=daily_path_var, wraplength=580).pack()

tk.Button(root, text="Select Historical Stats.xlsx", command=browse_historical).pack(pady=(10, 0))
tk.Label(root, textvariable=historical_path_var, wraplength=580).pack()

tk.Label(root, text="Enter Number of Recent Rows to Process:").pack(pady=(10, 0))
count_entry = tk.Entry(root)
count_entry.pack()

tk.Button(root, text="Run", command=run_conversion, bg="green", fg="white").pack(pady=20)

result_label = tk.Label(root, text="", font=("Arial", 10))
result_label.pack()

root.mainloop()
