import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook
from collections import defaultdict

def get_column_n_values_by_name(daily_path, name):
    wb = load_workbook(daily_path, data_only=True)
    ws = wb.active
    results = []
    for row in ws.iter_rows(min_row=1):
        if row[0].value == name:
            results.append(row[13].value)  # Column N
    return results

def get_filtered_rows(historical_path, symbol, count):
    wb = load_workbook(historical_path, data_only=True)
    ws = wb.active

    found = False
    max_count = int(count)
    data_collected = 0
    match_rows = []
    c_count = 0

    for i, row in enumerate(ws.iter_rows(min_row=1), start=1):
        if found:
            if all(cell.value is None for cell in row):
                break
            match_rows.append(row)
        elif row[1].value == symbol:
            found = True

    for row in reversed(match_rows):
        b_val = row[1].value
        c_val = row[2].value
        if isinstance(b_val, (int, float)):
            c_count += b_val
        if isinstance(c_val, (int, float)):
            c_count -= c_val
        data_collected += 1
        if data_collected >= max_count:
            break

    return c_count

def process_daily_format(daily_format_path, daily_path, historical_path, count_val, start_val, end_val, output_dir, prefix):
    wb = load_workbook(daily_format_path)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(min_row=1), start=1):
        a_value = row[0].value
        symbols = get_column_n_values_by_name(daily_path, a_value)
        t_count = 0
        for symbol in symbols:
            temp_count = get_filtered_rows(historical_path, symbol, count_val)
            if start_val is not None and end_val is not None:
                if (start_val <= temp_count <= end_val) or (-end_val <= temp_count <= -start_val):
                    t_count += temp_count
            else:
                t_count += temp_count
        row[8].value = t_count
        print(f"Row {i} - t_count: {t_count}")
    
    os.makedirs(output_dir, exist_ok=True)
    out_file = os.path.join(output_dir, f"{prefix}Modified_Daily_Format.xlsx")
    wb.save(out_file)
    return out_file

def extract_prefix(filename):
    return filename.split('_')[0] + '_' if '_' in filename else ''

def run_batch_process():
    d_fmt_dir = daily_format_dir.get()
    d_dir = daily_dir.get()
    h_dir = historical_dir.get()
    count_val = count_entry.get()
    start_val = start_entry.get()
    end_val = end_entry.get()

    if not d_fmt_dir or not d_dir or not h_dir or not count_val:
        result_label.config(text="Please enter all folders and Count value.", fg="red")
        return

    try:
        count_val = int(count_val)
        start_val = int(start_val) if start_val else None
        end_val = int(end_val) if end_val else None
    except ValueError:
        result_label.config(text="Invalid numeric value.", fg="red")
        return

    result_label.config(text="Processing...", fg="blue")
    root.update()

    fmt_files = {extract_prefix(f): os.path.join(d_fmt_dir, f) for f in os.listdir(d_fmt_dir) if f.endswith('.xlsx')}
    daily_files = {extract_prefix(f): os.path.join(d_dir, f) for f in os.listdir(d_dir) if f.endswith('.xlsx')}
    historical_files = {extract_prefix(f): os.path.join(h_dir, f) for f in os.listdir(h_dir) if f.endswith('.xlsx')}

    common_prefixes = set(fmt_files.keys()) & set(daily_files.keys()) & set(historical_files.keys())
    success = 0
    for prefix in common_prefixes:
        print(prefix)
        try:
            output_dir = os.path.join(os.path.dirname(d_fmt_dir), "Output")
            process_daily_format(
                fmt_files[prefix], daily_files[prefix], historical_files[prefix],
                count_val, start_val, end_val,
                output_dir, prefix
            )
            success += 1
        except Exception as e:
            print(f"[{prefix}] Error during processing: {e}")
            continue

    result_label.config(text=f"Processed {success} file sets.", fg="green")

# --- GUI ---
root = tk.Tk()
root.title("Batch Daily_Format Processor")
root.geometry("600x450")

daily_format_dir = tk.StringVar()
daily_dir = tk.StringVar()
historical_dir = tk.StringVar()

def browse_folder(var, title):
    folder = filedialog.askdirectory(title=title)
    if folder:
        var.set(folder)

tk.Button(root, text="Select Daily_Format Directory", command=lambda: browse_folder(daily_format_dir, "Select Daily_Format Folder")).pack(pady=5)
tk.Label(root, textvariable=daily_format_dir, wraplength=580).pack()

tk.Button(root, text="Select Daily Directory", command=lambda: browse_folder(daily_dir, "Select Daily Folder")).pack(pady=5)
tk.Label(root, textvariable=daily_dir, wraplength=580).pack()

tk.Button(root, text="Select Historical Directory", command=lambda: browse_folder(historical_dir, "Select Historical Folder")).pack(pady=5)
tk.Label(root, textvariable=historical_dir, wraplength=580).pack()

tk.Label(root, text="How many recent rows to process?").pack(pady=(10, 0))
count_entry = tk.Entry(root)
count_entry.pack()

tk.Label(root, text="Range (optional)").pack(pady=(10, 0))
range_frame = tk.Frame(root)
range_frame.pack()

start_entry = tk.Entry(range_frame, width=10)
start_entry.grid(row=0, column=0, padx=(0, 5))
tk.Label(range_frame, text="~").grid(row=0, column=1)
end_entry = tk.Entry(range_frame, width=10)
end_entry.grid(row=0, column=2, padx=(5, 0))

tk.Button(root, text="Start Processing", command=run_batch_process, bg="green", fg="white").pack(pady=20)

result_label = tk.Label(root, text="", font=("Arial", 11))
result_label.pack()

root.mainloop()
