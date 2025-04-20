import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook

def get_column_n_values_by_name(daily_path, name):
    wb = load_workbook(daily_path, data_only=True)
    ws = wb.active

    results = []
    for row in ws.iter_rows(min_row=1):  # Iterate through all rows
        a_cell = row[0]  # Column A (index 0)
        if a_cell.value == name:
            n_cell = row[13]  # Column N (index 13)
            results.append(n_cell.value)

    return results

def get_filtered_rows(historical_path, symbol, count):
    wb = load_workbook(historical_path, data_only=True)
    ws = wb.active

    found = False
    max_count = int(count)
    data_collected = 0
    match_rows = []
    c_count = 0

    for i, row in enumerate(ws.iter_rows(min_row=1), start=1):
        b_cell = row[1]  # Column B

        if found:
            if all(cell.value is None for cell in row):
                break  # Stop when reaching an empty row

            try:
                match_rows.append(row)
            except Exception as e:
                print(f"Error parsing row {i}: {e}")
                continue

        elif b_cell.value == symbol:
            found = True  # Start from the matched symbol

    for i in reversed(range(len(match_rows))):
        row = match_rows[i]

        b_val = row[1].value  # Column B (duplicate with symbol, can be omitted)
        c_val = row[2].value  # Column C

        if isinstance(b_val, (int, float)):
            c_count += b_val

        if isinstance(c_val, (int, float)):
            c_count -= c_val

        data_collected += 1
        if data_collected >= max_count:
            break

    return c_count

def process_daily_format(daily_format_path, daily_path_path, historical_path, count_val, start_val=None, end_val=None):
    wb = load_workbook(daily_format_path)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(min_row=1), start=1):
        a_value = row[0].value  # Column A
        symbols = get_column_n_values_by_name(daily_path_path, a_value)
        
        t_count = 0
        for symbol in symbols:
            temp_count = get_filtered_rows(historical_path, symbol, count_val)
            if start_val and end_val:
                # Check if temp_count is within positive range (start_val to end_val)
                # or negative range (-start_val to -end_val)
                if (start_val <= temp_count <= end_val) or (-end_val <= temp_count <= -start_val):
                    t_count += temp_count
            else:
                t_count += temp_count


        print(f"Row {i} - t_count: {t_count}")
        row[8].value = t_count  # I column (index 8)

    out_path = os.path.join(os.path.dirname(daily_format_path), "Modified_Daily_Format.xlsx")
    wb.save(out_path)

    result_label.config(text=f"Completed!\nSaved to {out_path}", fg="green")

def browse_daily_format():
    path = filedialog.askopenfilename(title="Select Daily_Format.xlsx", filetypes=[("Excel files", "*.xlsx")])
    if path:
        daily_format_path_var.set(path)

def browse_daily():
    path = filedialog.askopenfilename(title="Select Daily.xlsx", filetypes=[("Excel files", "*.xlsx")])
    if path:
        daily_path_var.set(path)

def browse_historical():
    path = filedialog.askopenfilename(title="Select Historical Stats.xlsx", filetypes=[("Excel files", "*.xlsx")])
    if path:
        historical_path_var.set(path)

def run_conversion():
    daily_format_path = daily_format_path_var.get()
    daily_path_path = daily_path_var.get()
    historical_path = historical_path_var.get()
    count_val = count_entry.get()
    start_val = start_entry.get()
    end_val = end_entry.get()

    if not daily_format_path or not daily_path_path or not historical_path or not count_val:
        result_label.config(text="Please select all files and input a count value.", fg="red")
        return
    
    try:
        count_val = int(count_val)
        start_val = int(start_val) if start_val else None
        end_val = int(end_val) if end_val else None
    except ValueError:
        result_label.config(text="Start/End, Count must be integers if provided.", fg="red")
        return

    result_label.config(text="Processing...", fg="blue")
    root.update()

    try:
        process_daily_format(daily_format_path, daily_path_path, historical_path, count_val, start_val, end_val)
    except Exception as e:
        print(e)
        result_label.config(text=f"Error: {str(e)}", fg="red")

# --- GUI SETUP ---
root = tk.Tk()
root.title("Daily Format Processor")
root.geometry("600x400")

daily_format_path_var = tk.StringVar()
daily_path_var = tk.StringVar()
historical_path_var = tk.StringVar()

tk.Button(root, text="Select Daily_Format.xlsx", command=browse_daily_format).pack(pady=(10, 0))
tk.Label(root, textvariable=daily_format_path_var, wraplength=580).pack()

tk.Button(root, text="Select Daily.xlsx", command=browse_daily).pack(pady=(10, 0))
tk.Label(root, textvariable=daily_path_var, wraplength=580).pack()

tk.Button(root, text="Select Historical Stats.xlsx", command=browse_historical).pack(pady=(10, 0))
tk.Label(root, textvariable=historical_path_var, wraplength=580).pack()

tk.Label(root, text="Enter Number of Recent Rows to Process:").pack(pady=(10, 0))
count_entry = tk.Entry(root)
count_entry.pack()

range_frame = tk.Frame(root)
range_frame.pack(pady=(10, 0))

tk.Label(range_frame, text="Range (Optional):").grid(row=0, column=0, padx=(0, 5))
start_entry = tk.Entry(range_frame, width=10)
start_entry.grid(row=0, column=1, padx=(0, 15))

tk.Label(range_frame, text="~ ").grid(row=0, column=2, padx=(0, 5))
end_entry = tk.Entry(range_frame, width=10)
end_entry.grid(row=0, column=3)
tk.Button(root, text="Run", command=run_conversion, bg="green", fg="white").pack(pady=20)

result_label = tk.Label(root, text="", font=("Arial", 10))
result_label.pack()

root.mainloop()
