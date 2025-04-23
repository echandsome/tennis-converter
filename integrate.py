import pandas as pd
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import platform
import openpyxl
import tempfile
from openpyxl import load_workbook, Workbook
import re
import os
import csv
import cv2
import copy
import json
import time
import numpy
import random
import codecs
import shutil
import urllib
import requests
from lxml.html import fromstring

TEMP_DIR = tempfile.mkdtemp()

# Month name mapping
MONTH_MAP = {
    '01': 'Jan', '02': 'Feb', '03': 'Mar', '04': 'Apr',
    '05': 'May', '06': 'Jun', '07': 'Jul', '08': 'Aug',
    '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dec'
}

# Function to separate the date
def format_date(date_str):
    if pd.isna(date_str):
        return ['', '', '']
    try:
        date = pd.to_datetime(date_str)
        return [str(date.day).zfill(2), MONTH_MAP[str(date.month).zfill(2)], str(date.year)]
    except:
        return ['', '', '']
    
def convert_date_format(date_input):
    if isinstance(date_input, pd.Timestamp):
        dt = date_input.to_pydatetime()
    elif isinstance(date_input, datetime):
        dt = date_input
    elif isinstance(date_input, str):
        dt = datetime.strptime(date_input, "%Y-%m-%d %H:%M:%S")
    else:
        raise ValueError("Input must be a string, datetime, or pandas.Timestamp")

    if platform.system() == "Windows":
        return dt.strftime("%#m/%#d/%Y")
    else:
        return dt.strftime("%-m/%-d/%Y")

def convert_date(date_str):
    """Convert 'DD_MMM_YYYY' to 'M/D/YYYY' format."""
    if pd.isna(date_str):
        return None
    
    clean_date = re.sub(r"[()]", "", date_str.strip())

    parts = clean_date.split("_")
    if len(parts) == 3:
        day, month, year = parts
        try:
            converted_date = datetime.strptime(f"{day} {month} {year}", "%d %b %Y").strftime("%m/%d/%Y")
            return converted_date.lstrip("0").replace("/0", "/")
        except ValueError:
            return None
    return None


def Step1_Tennis_Matches_Daily_Present_into_Astro(input_path):
    df = pd.read_excel(input_path, header=0)

    output_rows = []

    for _, row in df.iterrows():
        if pd.isna(row.iloc[0]) and pd.isna(row.iloc[3]) and pd.isna(row.iloc[14]):
            continue  # Remove completely empty rows

        # Date
        date_parts = format_date(row.iloc[0])
        # Location
        location = row.iloc[3] if not pd.isna(row.iloc[3]) else ''
        # O~R columns (index 14~17)
        extra_cols = [
            int(row.iloc[14]) if not pd.isna(row.iloc[14]) else '01',
            row.iloc[15] if not pd.isna(row.iloc[15]) else 'Jan',
            int(row.iloc[16]) if not pd.isna(row.iloc[16]) else '2001',
            row.iloc[17] if not pd.isna(row.iloc[17]) else 'Moscow, Russia'
        ]

        row_data = date_parts + [location] + [''] * 3 + extra_cols
        output_rows.append(row_data)

    # Keep 2 header rows
    header1 = ['Partner A', 'unknown time ON', '', '', '', '', '', 'Partner B', 'unknown time ON', '', '']
    header2 = ['Day', 'Month', 'Year', 'Location', '', '', '', 'Day', 'Month', 'Year', 'Location']
    output_data = [header1, header2] + output_rows

    out_dir = os.path.dirname(input_path)
    output_path_xlsx = os.path.join(out_dir, "Converted_Astro.xlsx")
    output_path_csv = os.path.join(out_dir, "Converted_Astro.csv")

    df_out = pd.DataFrame(output_data)

    # Save as Excel
    with pd.ExcelWriter(output_path_xlsx, engine='openpyxl') as writer:
        df_out.to_excel(writer, index=False, header=False)

    # Save as CSV (UTF-8 encoding)
    df_out.to_csv(output_path_csv, index=False, header=False, encoding='utf-8-sig')
    
    print(f"Step1 complete! Saved to:{os.path.basename(output_path_xlsx)} and {os.path.basename(output_path_csv)}")

    return 'Converted_Astro.csv'

class Step2_AstroCompatAnalyzer:
    def __init__(self, symbols_dir='Symbols', output_dir='ImageFiles', search_cache='SearchFiles',
                 output_csv='Astro_Output.csv'):
        print (f'\n ########### :: Horoscopes Images Data Extractor :: ############\n')

        if any(not os.path.isfile(f'{symbols_dir}/{img}.png') for img in ['BLK_Balls','BLK_Poker','BLK_Square','BLK_Star','BLK_Triangle']):
            raise Exception ('Failed to Location Template Images.')
        
        if not os.path.exists(search_cache):
            os.makedirs(search_cache)

        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        self.output_csv = output_csv

    def _SaveCSVData(self, row, mode='a'):
        try:
            with open(self.output_csv, mode, encoding='utf-8-sig', newline='') as csv_file:
                csv_writer = csv.writer(csv_file, quoting=csv.QUOTE_NONNUMERIC)
                csv_writer.writerow(row)
                
        except Exception as e:
            print (f'\n # CSV File Error: {e}\n')
            
    def _GetWebPage(self, url, path=None):

        response = None

        try:
            if not path or not os.path.isfile(path):

                time.sleep(random.randint(1,2))

                if '.png' in path:
                    response = self.http_session.get(url, stream=True, timeout=60)
                else:
                    response = self.http_session.get(url, timeout=60)

                if response.status_code != 200:
                    if response.status_code==404 or response.status_code==410:
                        return None

                    raise Exception (f'Response Code ({response.status_code})')
                    
                if '.png' in path:
                    response.raw.decode_content = True
                    response = response.raw

                    with open(path, 'wb') as image:
                        shutil.copyfileobj(response, image)
                else:
                    response = response.text
                    codecs.open(path,'w','utf-8').write(response)    
            else:
                if not '.png' in path:
                    response = codecs.open(path,'r','utf-8').read()
                
        except Exception as e:
            print (f'\n # HTTP Error: {e}\n')
            
            time.sleep(random.randint(2,3))

            return self._GetWebPage(url, path)

        return response   

    def _ExtractImageData(self, slug):

        img_org = cv2.imread(f'ImageFiles/{slug}.png')

        for template in ['BLK_Balls', 'BLK_Poker', 'BLK_Square', 'BLK_Star', 'BLK_Triangle']:

            try:
                img_rgb = copy.deepcopy(img_org)
                
                img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)

                img_template = cv2.imread(f'Symbols/{template}.png', cv2.IMREAD_GRAYSCALE)

                img_w, img_h = img_template.shape[::-1]
                
                result = cv2.matchTemplate(img_gray, img_template, cv2.TM_CCOEFF_NORMED)

                threshold, bounds = 0.95, {'x': 470, 'y': 425}

                markers = numpy.where(result >= threshold)

                for point in zip(*markers[::-1]):

                    if point[0] > bounds['x'] or point[1] > bounds['y']:
                        continue
                    
                    coord = (point[0] + img_w/2, point[1] + img_h/2)
                    
                    formula = lambda x, y: (x[0]-y[0])**2 + (x[1]-y[1])**2
                    
                    matched = min(self.lookup['coordinates'], key=lambda xy: formula(xy, coord))

                    label = self.lookup['indexes'][self.lookup['coordinates'].index(matched)]
                    
                    self._SaveCSVData(['', '', '', '', label, template])
                    
                    cv2.rectangle(img_rgb, point, (point[0] + img_w, point[1] + img_h), (0,0,255), 2)

                if not os.path.exists(f'ImageFiles/{slug}/'):
                    os.makedirs(f'ImageFiles/{slug}/')
                    
                cv2.imwrite(f'ImageFiles/{slug}/{template}.png', img_rgb)
            except:
                continue
        
    def _SearchLocation(self, slug):

        endpoint = f'https://horoscopes.astro-seek.com/api_gmaps3.php?term={slug}'
        
        for c in ['\\','/',':','*','?','"','<','>','|']:
            slug = slug.replace(c,'_')

        response = self._GetWebPage(endpoint, f'SearchFiles/{slug}.html')

        result = json.loads(response)

        endpoint = f"https://horoscopes.astro-seek.com/api_gmaps3.php?place_id={result[0]['id']}"

        response = self._GetWebPage(endpoint, f'SearchFiles/{slug}.json')
        
        details = json.loads(response)

        details.update({'id': result[0]['id'], 'city': result[0]['value']})

        return details
        
    def _SearchData(self, row):

        location_pa = self._SearchLocation(row[3])
        
        location_pb = self._SearchLocation(row[7])

        search_params = {
            'send_calculation': 1,
            'muz_narozeni_den': row[0],
            'muz_narozeni_mesic': row[1],
            'muz_narozeni_rok': row[2],
            'muz_narozeni_hodina': '00',
            'muz_narozeni_minuta': '00',
            'muz_narozeni_no_cas': 'on',
            'muz_narozeni_city': location_pa['city'],
            'muz_narozeni_mesto_hidden': location_pa['mesto'],
            'muz_narozeni_stat_hidden': location_pa['stat_kratky'],
            'muz_narozeni_podstat_kratky_hidden': location_pa['podstat_kratky'],
            'muz_narozeni_sirka_stupne': location_pa['sirka_stupne'],
            'muz_narozeni_sirka_minuty': location_pa['sirka_minuty'],
            'muz_narozeni_sirka_smer': location_pa['sirka_smer'],
            'muz_narozeni_delka_stupne': location_pa['delka_stupne'],
            'muz_narozeni_delka_minuty': location_pa['delka_minuty'],
            'muz_narozeni_delka_smer': location_pa['delka_smer'],
            'muz_narozeni_timezone_form': 'auto',
            'muz_narozeni_timezone_dst_form': 'auto',
            'send_calculation': 1,
            'zena_narozeni_den': row[4],
            'zena_narozeni_mesic': row[5],
            'zena_narozeni_rok': row[6],
            'zena_narozeni_hodina': '00',
            'zena_narozeni_minuta': '00',
            'zena_narozeni_no_cas': 'on',
            'zena_narozeni_city': location_pb['city'],
            'zena_narozeni_mesto_hidden': location_pb['mesto'],
            'zena_narozeni_stat_hidden': location_pb['stat_kratky'],
            'zena_narozeni_podstat_kratky_hidden': location_pa['podstat_kratky'],
            'zena_narozeni_sirka_stupne': location_pb['sirka_stupne'],
            'zena_narozeni_sirka_minuty': location_pb['sirka_minuty'],
            'zena_narozeni_sirka_smer': location_pb['sirka_smer'],
            'zena_narozeni_delka_stupne': location_pb['delka_stupne'],
            'zena_narozeni_delka_minuty': location_pb['delka_minuty'],
            'zena_narozeni_delka_smer': location_pb['delka_smer'],
            'zena_narozeni_timezone_form': 'auto',
            'zena_narozeni_timezone_dst_form': 'auto',
            'switch_interpretations': 3,
            'house_system': 'placidus',
            'hid_fortune': 1,
            'hid_fortune_check': 'on',
            'hid_vertex': 1,
            'hid_vertex_check': 'on',
            'hid_chiron': 1,
            'hid_chiron_check': 'on',
            'hid_lilith': 1,
            'hid_lilith_check': 'on',
            'hid_uzel': 1,
            'hid_uzel_check': 'on',
            'uhel_orbis': '',
            'hide_aspects': 0,
            'zmena_nastaveni': 1,
            'aktivni_tab': '',
            }

        slug = '-'.join([str(c) for c in [row[0], row[1], row[2], location_pa['id'], row[4], row[5], row[6], location_pb['id']]])

        endpoint = 'https://horoscopes.astro-seek.com/calculate-love-compatibility/?' + urllib.parse.urlencode(search_params)

        response = self._GetWebPage(endpoint, f'SearchFiles/{slug}.html')

        html = fromstring(response)

        partners = []

        for partner in html.xpath('//div[@class="detail-rozbor-items"]//div[contains(@style,"float: left; width: 250px;")]'):

            if not len(partner.xpath('.//em'))==2:
                continue
            
            dob = partner.xpath('.//em')[0].text_content().split('(')[0].strip().replace(' ','_')
            city = partner.xpath('.//em')[1].text_content().strip()
            
            if partner.xpath('.//strong[contains(text(),"Partner")]'):
                partners += [f'({dob})', city]

        if not len(partners)==4:
            return None

        print (' #', partners)
        
        self._SaveCSVData(partners)
        
        endpoint = html.xpath('//div[@id="tab4"]//img[@title="Aspect Tables"][@src]')[0].get('src')
        
        self._GetWebPage(endpoint, f'ImageFiles/{slug}.png')
        
        self._ExtractImageData(slug)

    def _start(self, input_csv='Astro_Input.csv'): 
        try:
           
            self.http_session = requests.Session()
            self.http_session.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36'})

            self.lookup = {}
            self.lookup['indexes'] = ['Sun_Sun', 'Moon_Sun', 'Mercury_Sun', 'Venus_Sun', 'Mars_Sun', 'Jupiter_Sun', 'Saturn_Sun', 'Uranus_Sun', 'Neptune_Sun', 'Pluto_Sun', 'Node_Sun', 'Lilith_Sun', 'Chiron_Sun', 'Sun_Moon', 'Moon_Moon', 'Mercury_Moon', 'Venus_Moon', 'Mars_Moon', 'Jupiter_Moon', 'Saturn_Moon', 'Uranus_Moon', 'Neptune_Moon', 'Pluto_Moon', 'Node_Moon', 'Lilith_Moon', 'Chiron_Moon', 'Sun_Mercury', 'Moon_Mercury', 'Mercury_Mercury', 'Venus_Mercury', 'Mars_Mercury', 'Jupiter_Mercury', 'Saturn_Mercury', 'Uranus_Mercury', 'Neptune_Mercury', 'Pluto_Mercury', 'Node_Mercury', 'Lilith_Mercury', 'Chiron_Mercury', 'Sun_Venus', 'Moon_Venus', 'Mercury_Venus', 'Venus_Venus', 'Mars_Venus', 'Jupiter_Venus', 'Saturn_Venus', 'Uranus_Venus', 'Neptune_Venus', 'Pluto_Venus', 'Node_Venus', 'Lilith_Venus', 'Chiron_Venus', 'Sun_Mars', 'Moon_Mars', 'Mercury_Mars', 'Venus_Mars', 'Mars_Mars', 'Jupiter_Mars', 'Saturn_Mars', 'Uranus_Mars', 'Neptune_Mars', 'Pluto_Mars', 'Node_Mars', 'Lilith_Mars', 'Chiron_Mars', 'Sun_Jupiter', 'Moon_Jupiter', 'Mercury_Jupiter', 'Venus_Jupiter', 'Mars_Jupiter', 'Jupiter_Jupiter', 'Saturn_Jupiter', 'Uranus_Jupiter', 'Neptune_Jupiter', 'Pluto_Jupiter', 'Node_Jupiter', 'Lilith_Jupiter', 'Chiron_Jupiter', 'Sun_Saturn', 'Moon_Saturn', 'Mercury_Saturn', 'Venus_Saturn', 'Mars_Saturn', 'Jupiter_Saturn', 'Saturn_Saturn', 'Uranus_Saturn', 'Neptune_Saturn', 'Pluto_Saturn', 'Node_Saturn', 'Lilith_Saturn', 'Chiron_Saturn', 'Sun_Uranus', 'Moon_Uranus', 'Mercury_Uranus', 'Venus_Uranus', 'Mars_Uranus', 'Jupiter_Uranus', 'Saturn_Uranus', 'Uranus_Uranus', 'Neptune_Uranus', 'Pluto_Uranus', 'Node_Uranus', 'Lilith_Uranus', 'Chiron_Uranus', 'Sun_Neptune', 'Moon_Neptune', 'Mercury_Neptune', 'Venus_Neptune', 'Mars_Neptune', 'Jupiter_Neptune', 'Saturn_Neptune', 'Uranus_Neptune', 'Neptune_Neptune', 'Pluto_Neptune', 'Node_Neptune', 'Lilith_Neptune', 'Chiron_Neptune', 'Sun_Pluto', 'Moon_Pluto', 'Mercury_Pluto', 'Venus_Pluto', 'Mars_Pluto', 'Jupiter_Pluto', 'Saturn_Pluto', 'Uranus_Pluto', 'Neptune_Pluto', 'Pluto_Pluto', 'Node_Pluto', 'Lilith_Pluto', 'Chiron_Pluto', 'Sun_Node', 'Moon_Node', 'Mercury_Node', 'Venus_Node', 'Mars_Node', 'Jupiter_Node', 'Saturn_Node', 'Uranus_Node', 'Neptune_Node', 'Pluto_Node', 'Node_Node', 'Lilith_Node', 'Chiron_Node', 'Sun_Lilith', 'Moon_Lilith', 'Mercury_Lilith', 'Venus_Lilith', 'Mars_Lilith', 'Jupiter_Lilith', 'Saturn_Lilith', 'Uranus_Lilith', 'Neptune_Lilith', 'Pluto_Lilith', 'Node_Lilith', 'Lilith_Lilith', 'Chiron_Lilith', 'Sun_Chiron', 'Moon_Chiron', 'Mercury_Chiron', 'Venus_Chiron', 'Mars_Chiron', 'Jupiter_Chiron', 'Saturn_Chiron', 'Uranus_Chiron', 'Neptune_Chiron', 'Pluto_Chiron', 'Node_Chiron', 'Lilith_Chiron', 'Chiron_Chiron']
            self.lookup['coordinates'] = [(155.0, 110.0), (155.0, 135.0), (155.0, 160.0), (155.0, 185.0), (155.0, 210.0), (155.0, 235.0), (155.0, 260.0), (155.0, 285.0), (155.0, 310.0), (155.0, 335.0), (155.0, 360.0), (155.0, 385.0), (155.0, 410.0), (180.0, 110.0), (180.0, 135.0), (180.0, 160.0), (180.0, 185.0), (180.0, 210.0), (180.0, 235.0), (180.0, 260.0), (180.0, 285.0), (180.0, 310.0), (180.0, 335.0), (180.0, 360.0), (180.0, 385.0), (180.0, 410.0), (205.0, 110.0), (205.0, 135.0), (205.0, 160.0), (205.0, 185.0), (205.0, 210.0), (205.0, 235.0), (205.0, 260.0), (205.0, 285.0), (205.0, 310.0), (205.0, 335.0), (205.0, 360.0), (205.0, 385.0), (205.0, 410.0), (230.0, 110.0), (230.0, 135.0), (230.0, 160.0), (230.0, 185.0), (230.0, 210.0), (230.0, 235.0), (230.0, 260.0), (230.0, 285.0), (230.0, 310.0), (230.0, 335.0), (230.0, 360.0), (230.0, 385.0), (230.0, 410.0), (255.0, 110.0), (255.0, 135.0), (255.0, 160.0), (255.0, 185.0), (255.0, 210.0), (255.0, 235.0), (255.0, 260.0), (255.0, 285.0), (255.0, 310.0), (255.0, 335.0), (255.0, 360.0), (255.0, 385.0), (255.0, 410.0), (280.0, 110.0), (280.0, 135.0), (280.0, 160.0), (280.0, 185.0), (280.0, 210.0), (280.0, 235.0), (280.0, 260.0), (280.0, 285.0), (280.0, 310.0), (280.0, 335.0), (280.0, 360.0), (280.0, 385.0), (280.0, 410.0), (305.0, 110.0), (305.0, 135.0), (305.0, 160.0), (305.0, 185.0), (305.0, 210.0), (305.0, 235.0), (305.0, 260.0), (305.0, 285.0), (305.0, 310.0), (305.0, 335.0), (305.0, 360.0), (305.0, 385.0), (305.0, 410.0), (330.0, 110.0), (330.0, 135.0), (330.0, 160.0), (330.0, 185.0), (330.0, 210.0), (330.0, 235.0), (330.0, 260.0), (330.0, 285.0), (330.0, 310.0), (330.0, 335.0), (330.0, 360.0), (330.0, 385.0), (330.0, 410.0), (355.0, 110.0), (355.0, 135.0), (355.0, 160.0), (355.0, 185.0), (355.0, 210.0), (355.0, 235.0), (355.0, 260.0), (355.0, 285.0), (355.0, 310.0), (355.0, 335.0), (355.0, 360.0), (355.0, 385.0), (355.0, 410.0), (380.0, 110.0), (380.0, 135.0), (380.0, 160.0), (380.0, 185.0), (380.0, 210.0), (380.0, 235.0), (380.0, 260.0), (380.0, 285.0), (380.0, 310.0), (380.0, 335.0), (380.0, 360.0), (380.0, 385.0), (380.0, 410.0), (405.0, 110.0), (405.0, 135.0), (405.0, 160.0), (405.0, 185.0), (405.0, 210.0), (405.0, 235.0), (405.0, 260.0), (405.0, 285.0), (405.0, 310.0), (405.0, 335.0), (405.0, 360.0), (405.0, 385.0), (405.0, 410.0), (430.0, 110.0), (430.0, 135.0), (430.0, 160.0), (430.0, 185.0), (430.0, 210.0), (430.0, 235.0), (430.0, 260.0), (430.0, 285.0), (430.0, 310.0), (430.0, 335.0), (430.0, 360.0), (430.0, 385.0), (430.0, 410.0), (455.0, 110.0), (455.0, 135.0), (455.0, 160.0), (455.0, 185.0), (455.0, 210.0), (455.0, 235.0), (455.0, 260.0), (455.0, 285.0), (455.0, 310.0), (455.0, 335.0), (455.0, 360.0), (455.0, 385.0), (455.0, 410.0)]

            with open(input_csv, 'r', encoding='utf-8-sig') as csv_file:
                csv_reader = csv.reader(csv_file)

                self._SaveCSVData(['Partner A', '', 'Partner B', '', 'Signs', 'Symbol'], 'w')

                for row in csv_reader:
                    
                    try:
                        row = [item.strip() for item in row if item.strip()]

                        if not len(row)==8:
                            raise Exception
                        
                        row = [int(row[0]), datetime.strptime(row[1], '%b').month, int(row[2]), row[3], int(row[4]), datetime.strptime(row[5], '%b').month, int(row[6]), row[7]]

                        self._SearchData(row)
                    except:
                        continue
                    
        except Exception as e:
            print (f'\n # Step2 Fatal Error: {e}\n')
            input ('\n # Press any Key to Exit..')

        print (f'\n ###############################################################')

        print(f"\nStep2 complete! Saved to: {self.output_csv}")

        return self.output_csv

def Step3_Tennis_Script_Pairs_WL_BP_File_Maker (matches_path, players_path):
    matches_df = pd.read_excel(matches_path, header=0)
    players_df = pd.read_excel(players_path, header=0)

    output_rows = []

    for _, match_row in matches_df.iterrows():
        player_name = match_row.iloc[2]  # Column C
        date = match_row.iloc[0]         # Column A
        result = match_row.iloc[18]      # Column S

        if pd.isna(player_name) or player_name.strip() == '':
            continue

        try:
            formatted_date = '"' + pd.to_datetime(date).strftime('%m-%d') + '"'
        except:
            formatted_date = '""'

        player_row = players_df[players_df['Player'] == player_name]
        if not player_row.empty:
            birth_place = player_row.iloc[0, 1]
            date_series = pd.to_datetime([player_row.iloc[0, 2]])
            date_of_birth = convert_date_format(date_series[0])
        else:
            birth_place = ''
            date_of_birth = ''

        row = [''] * 75
        row[52] = player_name

        for i in range(53, 59):
            row[i] = 1

        row[59] = formatted_date

        for i in range(60, 62):
            row[i] = 1

        row[62] = birth_place
        row[63] = date_of_birth

        for i in range(64, 74):
            row[i] = 1

        if result == 'W':
            row[74] = 'WIN'
        elif result == 'L':
            row[74] = 'LOSE'

        output_rows.append(row)

    header_temps = ["Player Name", "Number", "Odds", "Projection", "Avg", "Home/Away Avg", "Home/Away", "Date",
                    "Stat Category", "Team", "Place of Birth", "Date of Birth", "Illumination", "Age", "Type",
                    "Moon Cycle", "Result", "H/A DIF", "H/A Results DIF", "H/A Spread Result", "AVG DIF", "H / A DIF", "RESULT O/U"]

    headers = [""] * 52 + header_temps

    output_df = pd.DataFrame(output_rows, columns=headers)
    out_dir = os.path.dirname(matches_path)
    output_file = os.path.join(out_dir, "Tennis_Output_Example.xlsx")
    output_df.to_excel(output_file, index=False)

    # Auto-adjust column widths
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    for col in ws.columns:
        header_value = col[0].value
        col_letter = col[0].column_letter

        if not header_value or str(header_value).strip() == "":
            ws.column_dimensions[col_letter].width = 10
            continue

        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_file)

    print(f"Step3 complete! File saved as 'Tennis_Output_Example.xlsx'")

    return 'Tennis_Output_Example.xlsx'

def Step4_Tennis_Astro_Stats_Tennis (csv_file, ref_file):
    output_path = os.path.join(TEMP_DIR, "output.xlsx")
    ref_wb = load_workbook(ref_file, data_only=True)
    ref_ws = ref_wb.active
    df = pd.read_csv(csv_file, header=None)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Partner A", "", "Partner B", ""] + [""] * 7 + ["Signs", "Symbol"])
    
    ref_col_indices = [53, 64, 54, 69, 62, 73, 74, 75, 65, 67, 68]
    
    ref_row_num = 2
    
    for _, row in df.iterrows():
        if pd.notna(row[0]) and pd.notna(row[2]):
            dob_converted = convert_date(row[2]) if pd.notna(row[2]) else None
            ws.append([row[0], row[1], row[2], row[3]] + [""] * 7)
            if dob_converted:
                if ref_row_num <= ref_ws.max_row:
                    extracted_data = [
                        ref_ws.cell(row=ref_row_num, column=col).value for col in ref_col_indices
                    ]
                    ws.append(extracted_data)
                    ref_row_num += 1
        elif pd.notna(row[4]):
            ws.append([""] * 11 + [row[4], row[5]])
    
    wb.save(output_path)

    origin_file_name = os.path.basename(output_path)
    output_file_name = f'Result____{origin_file_name}.xlsx'

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    ws[f"N1"] = 'Signs-Symbol'
    
    for row in range(2, ws.max_row + 1):
        l_value = ws[f"L{row}"].value or ""
        m_value = ws[f"M{row}"].value or ""
        ws[f"N{row}"] = f"{l_value}-{m_value}" if l_value or m_value else ""

    for row in range(2, ws.max_row):  
        ws[f"L{row}"] = ws[f"L{row + 1}"].value
        ws[f"M{row}"] = ws[f"M{row + 1}"].value
        ws[f"N{row}"] = ws[f"N{row + 1}"].value

    ws[f"L{ws.max_row}"] = None
    ws[f"M{ws.max_row}"] = None
    ws[f"N{ws.max_row}"] = None

    columns_to_fill = [chr(col) for col in range(ord('A'), ord('K') + 1)]
    last_filled_values = {col: None for col in columns_to_fill}

    for row in range(2, ws.max_row):
        if all(ws[f"A{row}"].value is None for col in columns_to_fill) and ws[f"L{row}"].value is not None:
            for col in columns_to_fill:
                ws[f"{col}{row}"] = last_filled_values[col]
        else:
            for col in columns_to_fill:
                last_filled_values[col] = ws[f"{col}{row}"].value
    
    output_path = os.path.join(os.path.dirname(ref_file), output_file_name)
    wb.save(output_path)
    print(f"\nStep4 complete! Saved to: {output_path}")

# GUI - File selection
def browse_input():
    path = filedialog.askopenfilename(title="Select Input Excel File", filetypes=[("Excel files", "*.xlsx")])
    if path:
        input_path_var.set(path)

def browse_players():
    path = filedialog.askopenfilename(title="Select Input_Tennis_Players.xlsx", filetypes=[("Excel files", "*.xlsx")])
    if path:
        players_path_var.set(path)

# GUI - Run button
def run_conversion():
    input_path = input_path_var.get()
    players_path = players_path_var.get()

    if not input_path or not players_path:
        result_label.config(text="Please select both files first.", fg="red")
        return
    result_label.config(text="Processing...", fg="blue")
    root.update()
    try:
        astro_path = Step1_Tennis_Matches_Daily_Present_into_Astro(input_path)

        analyzer = Step2_AstroCompatAnalyzer()
        analyze_path = analyzer._start(astro_path)
        
        result_path = Step3_Tennis_Script_Pairs_WL_BP_File_Maker(input_path, players_path)

        Step4_Tennis_Astro_Stats_Tennis(analyze_path, result_path)

        result_label.config(text=f"Completed! ", fg="green")
    except Exception as e:
        print(e)
        result_label.config(text=f"Error: {str(e)}", fg="red")

# --- GUI SETUP ---
root = tk.Tk()
root.title("Astro Excel Converter")
root.geometry("500x220")

input_path_var = tk.StringVar()
players_path_var = tk.StringVar()

tk.Button(root, text="Select Mens Excel File", command=browse_input).pack(pady=(10, 0))
tk.Label(root, textvariable=input_path_var, wraplength=480).pack()

tk.Button(root, text="Select Player File", command=browse_players).pack(pady=(10, 0))
tk.Label(root, textvariable=players_path_var, wraplength=480).pack()

tk.Button(root, text="Convert", command=run_conversion, bg="green", fg="white").pack(pady=20)

result_label = tk.Label(root, text="", font=("Arial", 10))
result_label.pack()

root.mainloop()
