import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import traceback

# Function to classify color based on F value
def classify_color(f_value):
    try:
        if f_value >= 0.2:
            return "green"
        elif -0.19 <= f_value <= 0.19:
            return "grey"
        else:
            return "red"
    except Exception:
        return "unknown"

# Function to load team lineups in groups
def load_team_groups(lineup_path):
    with open(lineup_path, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f.readlines()]

    groups = []
    current_group = []

    for line in lines:
        if line == "":
            if current_group:
                groups.append(current_group)
                current_group = []
        else:
            current_group.append(line)

    if current_group:
        groups.append(current_group)

    return groups

# Sorting and saving the results
def sort_nba_by_color(nba_path, lineup_path):
    df = pd.read_excel(nba_path, header=None)
    wb = load_workbook(nba_path)
    ws = wb.active

    # Classify colors
    color_labels = []
    for i in range(len(df)):
        row_idx = i + 2
        try:
            f_value = ws[f"F{row_idx}"].value
            color = classify_color(f_value)
            color_labels.append(color)
        except:
            color_labels.append("unknown")

    df["Color"] = color_labels

    team_groups = load_team_groups(lineup_path)

    sorted_rows = []
    grey_rows = []

    for group in team_groups:
        for team in group:
            team_df = df[df[2] == team]
            green = team_df[team_df["Color"] == "green"]
            grey = team_df[team_df["Color"] == "grey"]
            red = team_df[team_df["Color"] == "red"]

            combined = pd.concat([green, red])
            sorted_rows.append(combined)
            grey_rows.append(grey)

        sorted_rows.append(pd.DataFrame([[]]))
        sorted_rows.append(pd.DataFrame([[]]))

    if grey_rows:
        sorted_rows.append(pd.DataFrame([[]]))
        grey_rows_df = pd.concat(grey_rows, ignore_index=True)
        sorted_rows.append(grey_rows_df)

    sorted_rows_df = pd.concat(sorted_rows, ignore_index=True) if sorted_rows else pd.DataFrame()

    # Create a new workbook
    new_wb = Workbook()
    new_ws = new_wb.active

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    red_fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")    # Light red
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")   # Light grey

    for i, row in sorted_rows_df.iterrows():
        if row.isnull().all():
            continue  # Skip empty rows

        for j, value in enumerate(row[:-1]):  # Excluding the Color column
            new_ws.cell(row=i + 1, column=j + 1, value=value)

        color = row["Color"]
        if color == "green":
            new_ws.cell(row=i + 1, column=4).fill = green_fill  # Column D
        elif color == "red":
            new_ws.cell(row=i + 1, column=5).fill = red_fill    # Column E
        elif color == "grey":
            new_ws.cell(row=i + 1, column=4).fill = grey_fill   # Column D
            new_ws.cell(row=i + 1, column=5).fill = grey_fill   # Column E
            new_ws.cell(row=i + 1, column=6).fill = grey_fill   # Column F

    base_name = os.path.splitext(os.path.basename(nba_path))[0]
    output_path = os.path.join(os.path.dirname(nba_path), f"{base_name}_sort.xlsx")
    new_wb.save(output_path)
    return output_path

# GUI part
def browse_nba():
    path = filedialog.askopenfilename(title="Select NBA.xlsx", filetypes=[("Excel files", "*.xlsx")])
    if path:
        nba_path_var.set(path)

def browse_lineup():
    path = filedialog.askopenfilename(title="Select Team_Lineup.txt", filetypes=[("Text files", "*.txt")])
    if path:
        lineup_path_var.set(path)

def run_process():
    nba_path = nba_path_var.get()
    lineup_path = lineup_path_var.get()

    if not nba_path or not lineup_path:
        result_label.config(text="Please select both files.", fg="red")
        return

    result_label.config(text="Processing...", fg="blue")
    root.update()

    try:
        output_path = sort_nba_by_color(nba_path, lineup_path)
        result_label.config(text=f"Completed!\nSaved to:\n{output_path}", fg="green")
    except Exception as e:
        print("An error occurred:")
        traceback.print_exc()
        result_label.config(text=f"Error: {str(e)}", fg="red")

# GUI initialization
root = tk.Tk()
root.title("NBA Group Sorter")
root.geometry("500x320")

nba_path_var = tk.StringVar()
lineup_path_var = tk.StringVar()

tk.Button(root, text="Select NBA.xlsx", command=browse_nba).pack(pady=(10, 0))
tk.Label(root, textvariable=nba_path_var, wraplength=480).pack()

tk.Button(root, text="Select Team_Lineup.txt", command=browse_lineup).pack(pady=(10, 0))
tk.Label(root, textvariable=lineup_path_var, wraplength=480).pack()

tk.Button(root, text="Sort and Export", command=run_process, bg="green", fg="white").pack(pady=20)

result_label = tk.Label(root, text="", font=("Arial", 10))
result_label.pack()

root.mainloop()
