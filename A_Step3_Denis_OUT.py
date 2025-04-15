import pandas as pd
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os

# Function to process the Excel file
def process_file():
    try:
        # Get the file path from the file dialog
        input_file = filedialog.askopenfilename(title="Select Input Excel File", filetypes=[("Excel Files", "*.xlsx;*.xls")])
        
        if not input_file:
            return
        
        # Get the base name of the input file (without the path and extension)
        input_filename = os.path.splitext(os.path.basename(input_file))[0]
        
        # Create the output file name by adding "_fix" before the extension
        output_file = f"{input_filename}_fix.xlsx"

        # Load the Excel file into a pandas DataFrame
        df = pd.read_excel(input_file)


        # Delete the first row (index 0)
        df = df.drop(index=0)

        # Filter rows where column Z contains "Result"
        column_z_index = 25  # Z is the 26th column, so its index is 25
        df = df[df.iloc[:, column_z_index] == 'Result']

        # Drop columns X, Y, Z (columns 24, 25, 26)
        columns_to_drop = [23, 24, 25]  # X=24, Y=25, Z=26
        df = df.drop(df.columns[columns_to_drop], axis=1)

        # Calculate column X as V - W (V is 22nd column, W is 23rd column)
        column_v_index = 21  # V is the 22nd column (index 21)
        column_w_index = 22  # W is the 23rd column (index 22)
        df['X'] = df.iloc[:, column_v_index] - df.iloc[:, column_w_index]

        # Round columns V, W, and X to 2 decimal places
        df.iloc[:, column_v_index] = df.iloc[:, column_v_index].round(2)
        df.iloc[:, column_w_index] = df.iloc[:, column_w_index].round(2)
        df['X'] = df['X'].round(2)

        # Rearrange columns and move columns B, F, G, V, W, X
        cols_bfg = df.iloc[:, [1, 5, 6]]
        cols_vwx = df.iloc[:, [21, 22, 23]]
        df = df.drop(df.columns[[1, 5, 6, 21, 22, 23]], axis=1)
        df.insert(4, 'V', cols_vwx.iloc[:, 0])
        df.insert(5, 'W', cols_vwx.iloc[:, 1])
        df.insert(6, 'X', cols_vwx.iloc[:, 2])
        df.insert(21, 'B', cols_bfg.iloc[:, 0])
        df.insert(22, 'F', cols_bfg.iloc[:, 1])
        df.insert(23, 'G', cols_bfg.iloc[:, 2])

        # Sort the DataFrame by column 'G' (new position)
        df = df.sort_values(by=df.columns[6], ascending=False)

        # Delete the contents of column I
        df.iloc[:, 8] = None  # Column I is at index 8 (9th column), so set it to None


        df.to_excel(output_file, index=False, header=False)

        # Open the saved file to apply additional formatting
        wb = load_workbook(output_file)
        ws = wb.active

        # Define colors
        green_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
        red_fill = PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # Initialize lists for green, red, and gray rows
        green_rows = []
        red_rows = []
        gray_rows = []

        # Apply conditions for coloring and sorting rows
        for row in range(1, len(df) + 2):  # Adjusting for header rows
            # Get the value in column G
            g_value = ws.cell(row=row, column=7).value

            # Skip if the value is None or NaN
            if g_value is None or pd.isna(g_value):
                continue

            # Condition 1: Color column E green if G > 0.2
            if g_value > 0.2:
                green_rows.append(row)  # Store green rows
                ws.cell(row=row, column=5).fill = green_fill  # Column E is 5
            # Condition 2: Color column E red if G <= -0.2
            elif g_value <= -0.2:
                red_rows.append(row)  # Store red rows
                ws.cell(row=row, column=6).fill = red_fill  # Column E is 5
            # Condition 3: Color columns E, F, G gray and move them to the bottom if -0.19 <= G <= 0.19
            elif -0.19 <= g_value <= 0.19:
                gray_rows.append(row)  # Store gray rows
                ws.cell(row=row, column=5).fill = gray_fill  # Column E is 5
                ws.cell(row=row, column=6).fill = gray_fill  # Column F is 6
                ws.cell(row=row, column=7).fill = gray_fill  # Column G is 7

        # Save the file with all formatting applied
        wb.save(output_file)

        # Show success message
        messagebox.showinfo("Success", f"File processed successfully! Output saved as {output_file}")

    except Exception as e:
        # Handle errors and display a message
        messagebox.showerror("Error", f"An error occurred: {e}")

# Create the main Tkinter window
root = tk.Tk()
root.title("Excel File Processor")
root.geometry("300x150")

# Create a button to start processing
btn_process = tk.Button(root, text="Browse and Process Excel File", command=process_file)
btn_process.pack(pady=50)

# Run the Tkinter event loop
root.mainloop()
